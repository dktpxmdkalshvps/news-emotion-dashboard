"""
exporter.py - 뉴스 감성 분석 결과 엑셀 추출 (강화 버전)
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from logger import get_logger

logger = get_logger('exporter')


class DataExporter:
    def __init__(self, keyword: str, output_dir: str = "output"):
        self.keyword = keyword
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"데이터 익스포터 초기화: {keyword}")

    def export(self, df: pd.DataFrame) -> str:
        if df.empty:
            logger.warning("빈 데이터프레임: 엑셀 저장 불가")
            return ""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.keyword}_감성분석_{timestamp}.xlsx"
        path = os.path.join(self.output_dir, filename)

        logger.info(f"엑셀 저장 시작: {filename}")
        
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                # Sheet 1: 전체 분석 결과
                df.to_excel(writer, sheet_name="분석결과", index=False)
                
                # Sheet 2: 긍정 뉴스
                pos_df = df[df["sentiment"] == "긍정"]
                pos_df.to_excel(writer, sheet_name="긍정뉴스", index=False)
                logger.debug(f"긍정뉴스: {len(pos_df)}건")
                
                # Sheet 3: 부정 뉴스
                neg_df = df[df["sentiment"] == "부정"]
                neg_df.to_excel(writer, sheet_name="부정뉴스", index=False)
                logger.debug(f"부정뉴스: {len(neg_df)}건")
                
                # Sheet 4: 중립 뉴스
                neu_df = df[df["sentiment"] == "중립"]
                neu_df.to_excel(writer, sheet_name="중립뉴스", index=False)
                logger.debug(f"중립뉴스: {len(neu_df)}건")
                
                # Sheet 5: 통계 분석
                stats_df = self._generate_statistics_sheet(df)
                stats_df.to_excel(writer, sheet_name="통계", index=False)
                
                # 엑셀 스타일 적용
                self._apply_styles(writer, path)
            
            logger.info(f"엑셀 저장 완료: {path}")
            return path
            
        except Exception as e:
            logger.error(f"엑셀 저장 실패: {e}", exc_info=True)
            return ""

    def _generate_statistics_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """통계 분석 시트 생성"""
        stats = []
        
        total = len(df)
        if total == 0:
            return pd.DataFrame({"항목": ["데이터 없음"], "값": [0]})
        
        sentiment_counts = df["sentiment"].value_counts()
        
        stats.append(["항목", "값"])
        stats.append(["총 뉴스 수", total])
        stats.append(["긍정", sentiment_counts.get("긍정", 0)])
        stats.append(["부정", sentiment_counts.get("부정", 0)])
        stats.append(["중립", sentiment_counts.get("중립", 0)])
        stats.append(["긍정 비율(%)", f"{sentiment_counts.get('긍정', 0)/total*100:.1f}"])
        stats.append(["부정 비율(%)", f"{sentiment_counts.get('부정', 0)/total*100:.1f}"])
        stats.append(["평균 점수", f"{df['score'].mean():.3f}"])
        stats.append(["최고 점수", f"{df['score'].max():.2f}"])
        stats.append(["최저 점수", f"{df['score'].min():.2f}"])
        
        if "confidence" in df.columns:
            stats.append(["평균 신뢰도", f"{df['confidence'].mean():.3f}"])
        
        if "source" in df.columns:
            stats.append(["수집 출처 수", df["source"].nunique()])
        
        stats.append(["분석 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        return pd.DataFrame(stats[1:], columns=stats[0])

    def _apply_styles(self, writer, path: str):
        """엑셀 스타일 적용 (색상, 폰트, 정렬)"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(path)
            
            # 색상 정의
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            pos_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            neg_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            neu_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 헤더 스타일
                for cell in ws[1]:
                    if cell.value:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border
                
                # 데이터 스타일
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        
                        # 감정 컬럼 색칠 (sentiment 컬럼이 존재하면)
                        if cell.column_letter == 'D' and cell.value:  # sentiment 컬럼 (일반적으로 4번째)
                            if cell.value == "긍정":
                                cell.fill = pos_fill
                            elif cell.value == "부정":
                                cell.fill = neg_fill
                            elif cell.value == "중립":
                                cell.fill = neu_fill
                
                # 컬럼 너비 조정
                for col in ws.columns:
                    max_length = 0
                    column = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
            
            wb.save(path)
            logger.debug("엑셀 스타일 적용 완료")
            
        except ImportError:
            logger.warning("openpyxl 스타일링 기능을 사용할 수 없습니다.")
        except Exception as e:
            logger.warning(f"엑셀 스타일 적용 실패: {e}")
